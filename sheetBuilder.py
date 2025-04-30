#### Created by David N Preiss

# create template.xlsx file

# open pool.xlsx file
# iterate through the rows of pool.xlsx
# for each row1
	# if col E is blank
		# create a new line in template
		# far left cell is F contents
		# below it is A contents
		# iterate through rows under
		# for each row2
			# if col F is blank and col E is not blank
				# next col of template row is col E contents
				# below it is A + "|" + D contents
			# else
				# row1 = row1 + row2 -1
				# break
	# else print row1

# close pool.xlsx
# save template.xlsx

############

###   --Gloabal Values
source_path = "Customer and pool.xlsx"
target_path = "raw_template.xlsx"
sourceRowStart = 1
sourceRowEnd = 1800

###   --Import Statements

import shutil
import os
import calendar
import subprocess

## System call
os.system("")

## Class of different styles
class style():
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    UNDERLINE = '\033[4m'
    RESET = '\033[0m'


try:
    import openpyxl
except ImportError as e:
    print(style.YELLOW + f"!--WARNING:{e}\nopenpyxl is not installed. Installing..." + style.RESET)
    subprocess.check_call(["pip", "install", "openpyxl"])
    print("Installation complete.")

## create template.xlsx file

targetWorkbook = openpyxl.Workbook()
targetWorksheet = targetWorkbook.active

targetWorksheet.title = "DavidSheet1"

data = [
    ["Name", "Age", "City"],
    ["Alice", 30, "new york"],
    ["bob", 25, "los Angeles"],
    ["charlie", 35, "chicago"]
    ]



# for row in data:
    # targetWorksheet.append(row)

# targetWorksheet.cell(row=5, column=2).value = 69


## open pool.xlsx file
sourceWorkbook = openpyxl.load_workbook(source_path)
sourceWorksheet = sourceWorkbook.active

failCounter = 0
targetRow = 1
targetCol = 1
## iterate through the rows of source
row1 = 0
while (row1 < sourceRowEnd):
    ## for each row1
    row1 = row1+1
	## if col E is blank
    Acontent = sourceWorksheet.cell(row=row1, column=1).value
    Econtent = sourceWorksheet.cell(row=row1, column=5).value
    if Econtent is None and Acontent is not None:
        print(f"found customer! {row1}")
        ## create a new line in template
        targetRow = targetRow + 2
        targetCol = 1
		## far left cell is F contents
        Fcontent = sourceWorksheet.cell(row=row1, column=6).value
        targetWorksheet.cell(row=targetRow, column=targetCol).value = Fcontent
		## below it is A contents
        targetWorksheet.cell(row=targetRow+1, column=targetCol).value = Acontent
		
		## iterate through rows under
		## for each row2
        for row2 in range(1,30):
            Acontent2 = sourceWorksheet.cell(row=row1+row2, column=1).value
            Dcontent2 = sourceWorksheet.cell(row=row1+row2, column=4).value
            Econtent2 = sourceWorksheet.cell(row=row1+row2, column=5).value
            Fcontent2 = sourceWorksheet.cell(row=row1+row2, column=6).value
			## if col F is blank and col E is not blank
            if Fcontent2 is None and Econtent2 is not None:
				## next col of template row is col E contents
                print(f"inside {row1} + {row2}")
                targetCol = targetCol+1
                targetWorksheet.cell(row=targetRow, column=targetCol).value = Econtent2
                
				## below it is A + "|" + D contents
                combo = Acontent2 + "|" + Dcontent2
                targetWorksheet.cell(row=targetRow+1, column=targetCol).value = combo
            else:
                row1 = row1 + row2 -1
                break
    else:
        failCounter+=1
        print(f"\t FAIL {failCounter} at row {row1}")
print(f"FAILS: {failCounter}")
## close source
sourceWorkbook.close()
## save target
targetWorkbook.save(target_path)
targetWorkbook.close()

print("DONE!")
## save AllCustomers.html in Customers Folder