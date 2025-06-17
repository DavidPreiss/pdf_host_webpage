#### Created by David N Preiss

## TO DO:
# All Customers Page
# All Pools Page

### TABLE OF CONTENTS
"""
# Intro
# Hard-coded Global Values
# Import Statements
# Function Definitions
# Pseudocode
# Main Code
# Outro
"""

### Intro
"""
# This file iterates through an xlsx file (CUSTOMERS_POOLS_XLSX)
# to navigate the web archive (WEB_ARCHIVE_PATH)
# To find pool pdfs and organize them into a folder structure
# and then generate html files that navigate that structure
# refer to the pseudocode for details
"""

### Hard-coded Global Values
if True:
    MAX_WEEKS = 52
    WEB_ARCHIVE_PATH = "../../Web Archive"
    CUSTOMERS_POOLS_XLSX = "Customer Database.xlsx"
    C_ID_COL = 1
    C_ID_START_ROW = 1
    C_ID_END_ROW = 1000
    P_ID_START_COL = 2
    P_ID_END_COL = 1000

### Import Statements
if True:
    import shutil
    import os
    import calendar
    import subprocess
    # System call
    os.system("")

    # Class of different styles
    class style():
        BLACK = '\033[30m'
        RED = '\033[31m'
        GREEN = '\033[32m'
        YELLOW = '\033[33m'
        BLUE = '\033[34m'
        MAGENTA = '\033[35m'
        CYAN = '\033[36m'
        WHITE = '\033[37m'
        UNDERLINE = '\033[4m'
        RESET = '\033[0m'
    
    try:
        from datetime import datetime
    except ImportError as e:
        print(style.RED + f"!--ERROR:{e}\ndatetime is not installed. Installing..." + style.RESET)
        subprocess.check_call(["pip", "install", "datetime"])
        print("Installation complete. You can now run the script.")
        
    try:
        import openpyxl
    except ImportError as e:
        print(style.RED + f"!--ERROR:{e}\nopenpyxl is not installed. Installing..." + style.RESET)
        subprocess.check_call(["pip", "install", "openpyxl"])
        print("Installation complete. You can now run the script.")

### Function Definitions
if True:
    def iterateStats(statsFolder_path):
        
        statsFile_path = statsFolder_path+"/"+"statsFile.xlsx"
        if not os.path.exists(statsFolder_path):
            print(f"creating statsFolder_path: '{statsFolder_path}'") # debug
            os.makedirs(statsFolder_path)
            from openpyxl import Workbook

            # Create a workbook
            statswb = Workbook()

            # Get the active worksheet or create a new sheet
            ws = statswb.active

            ws.cell(1,1).value = "Program Name"
            ws.cell(1,2).value = "Times Ran on this CPU"
            ws.cell(1,3).value = "First Ran on this CPU"
            ws.cell(1,4).value = "Last Ran on this CPU"
            
            ws.cell(3,1).value = "WebsiteBuilder.py"
            ws.cell(3,2).value = 1
            ws.cell(3,3).value = datetime.now()
            ws.cell(3,4).value = datetime.now()
            
            # Save the workbook to a file
            statswb.save(statsFile_path)
            print(f"created: '{statsFile_path}'") # debug
        else:
            
            statswb = openpyxl.load_workbook(statsFile_path)
            ws = statswb.active
            if ws.cell(3,2).value is None:
                ws.cell(3,1).value = "WebsiteBuilder.py"
                ws.cell(3,2).value = 1
                ws.cell(3,3).value = datetime.now()
            else:
                ws.cell(3,2).value = ws.cell(3,2).value+1
            ws.cell(3,4).value = datetime.now()
            statswb.save(statsFile_path)
        print("iterated stats")

### Pseudocode
"""
# Create Customers Folder

# Open CustomersPools excel document

# Create AllCustomers.html in Customers Folder

# iterate through the CustomerID column in the xlsx
# for each CustomerID
    # create a CustomerID folder in the Customers folder
    # create a CustomerID.html file in that CustomerID folder
    # iterate through the row of the CustomerID in the xlsx
    # for each PoolID
        # create a PoolID folder in the CustomerID folder
        # inside the PoolID folder
            # create a PoolID.html file
            # for each of the last 52 pdfs of PoolID
                # import and rename the pdf into the PoolID folder
                # add a link to the PoolID.html that opens PoolDate.pdf
            # save PoolID.html
        # add a link to the CustomerID.html that leads to PoolID.html
    # save CustomerID.html
    # add a link to AllCustomers.html that leads to CustomerID.html

# save AllCustomers.html in Customers Folder
# close CustomersPools excel document
"""

### Main Code
if True:
    
    # Change working directory to the script's directory
    script_dir = os.path.dirname(os.path.realpath(__file__))
    os.chdir(script_dir)
    
    # Open CustomersPools excel document
    workbook = openpyxl.load_workbook(CUSTOMERS_POOLS_XLSX)
    worksheet = workbook.active


    # Create Customers Folder
    output_dir = "./Customers"
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    os.makedirs(output_dir, exist_ok=True)
    # print(f"Created output_dir:{output_dir}") # debug
    # output_dir = os.path.abspath(output_dir) # debug


    # Create AllCustomers.html in Customers Folder
    all_cids_html_name = "AllCustomers.html"
    all_cids_html_path = output_dir + "/" + all_cids_html_name
    all_cids_html_content= f"<!DOCTYPE html>\n<html>\n<head>\n<title>All Customers</title>\n</head>\n<body>\n<h2>All Customers</h2>\n<ul>\n"
    
    # est Array of T_IDs
    ArrayT_ID = []
    # est 2D Array of T_ID_num, T_name
    tripleArray = []
    # iterate through the CustomerID column in the xlsx
    for count in range(C_ID_START_ROW, C_ID_END_ROW):
        row = 2*count+1
        CustomerID = worksheet.cell(row=row, column=C_ID_COL).value
        if CustomerID is None:
            break #reached the end of CustomerIDs
        CustomerID = CustomerID.strip()
        CustomerName = worksheet.cell(row=row+1, column=C_ID_COL).value
        print(f"\n CustomerID: {CustomerID}")
        print(f" CustomerName: {CustomerName}")
        extraString = ""
        
        
        """
        Special catch needs to happen for CustomerIDs that start with T
        aka T_IDs
        Because There's going to be multiple of them
        There's already a loop that catches and renames duplicate CustomerIDs
        but we need something that builds out a branch page for our T_IDs
        Each actual numbered T_ID gets its own leaf page
        with a link at the top that leads back to the branch page
        """
        
        #blank string that gets populated in the following if statement
        tString = ""
        
        # if CustomerID starts with T
        if CustomerID[0]=='T':
            # check if Array of T_IDs contains CustomerID
            if CustomerID not in ArrayT_ID:
                #if not, add new T_ID to T_ID array and tripleArray
                ArrayT_ID.append(CustomerID)
                tempArray = []
                tripleArray.append(tempArray)
                
                # make a folder for the branch page
                cidFolder = output_dir + "/"+CustomerID
                os.makedirs(cidFolder, exist_ok=True)
                # make a link to the branch page in the All Customers page
                all_cids_html_content += f'    <li><a target="bottom" href="{CustomerID}/{CustomerID}.html">{CustomerID}</a> Branch Page</li>\n'
            
            # add numbered T_ID to triple array
            tripleArray[ArrayT_ID.index(CustomerID)].append(CustomerName)
            
            # make sure the non-numbered CustomerID is saved for the branch page
            extraString = "_1"
            
            # create a link back to the branch page at the top of the leaf page
            tString = f'<a target="bottom" href="../{CustomerID}/{CustomerID}.html">{CustomerID}</a>'
           
        # create a CustomerID folder in the Customers folder
        cidFolder = output_dir + "/"+CustomerID
        
        """
        Some CustomerIDs are duplicate
        Ideally this wouldn't happen,
        but when it does it needs to be handled.
        Here I just number them after an underscore
        """
        
        # check to see if this CustomerID was already caught
        tempcount=0
        while os.path.exists(cidFolder + extraString):
            tempcount+=1
            extraString = "_"+str(tempcount)
        CustomerID=CustomerID+extraString
        cidFolder = output_dir + "/"+CustomerID
        os.makedirs(cidFolder, exist_ok=True)
        # print(f" Created cidFolder:{cidFolder}") # debug
        
        # create a CustomerID.html file in that CustomerID folder
        cid_html_name = CustomerID+".html"
        cid_html_path = cidFolder + "/" + cid_html_name
        cid_html_content= f"<!DOCTYPE html>\n<html>\n<head>\n<title>{CustomerID}</title>\n</head>\n<body>\n{tString}<h2>{CustomerID}</h2>\n<h3>{CustomerName}</h3>\n<ul>\n"
        
        # iterate through the row of the CustomerID in the xlsx
        for col in range(P_ID_START_COL, P_ID_END_COL):
            PoolID = worksheet.cell(row=row, column=col).value
            if PoolID is None:
                break
            PoolID = str(PoolID).strip()
            PoolName = worksheet.cell(row=row+1, column=col).value
            print(f"   {PoolID}: {PoolName}") #debug
            # for each PoolID
            # create a PoolID folder in the CustomerID folder
            pidFolder = cidFolder+"/"+PoolID
            os.makedirs(pidFolder, exist_ok=True)
            # print(f"\t Created pidFolder:{pidFolder}") # debug
            
            # create a PoolID.html file in that PoolID folder
            pid_html_name = PoolID+".html"
            pid_html_path = pidFolder + "/" + pid_html_name
            pid_html_content= f"<!DOCTYPE html>\n<html>\n<head>\n<title>{PoolID}</title>\n</head>\n<body>\n<h2>{PoolID}</h2>\n<h3>{PoolName}</h3>\n<ul>\n"
            # iterate through the PoolID folder in the web archive
            WebArchive_PoolID_path = WEB_ARCHIVE_PATH+"/P"+str(PoolID)
            if os.path.exists(WebArchive_PoolID_path):
                fnames = os.listdir(WebArchive_PoolID_path)
                for file in range(0,MAX_WEEKS):
                    try:
                        fileName = os.path.basename(fnames[file])
                        # print(f"\t{fileName}") #debug
                        # import and rename the pdf into the PoolID folder
                        # Source path
                        src_path = WebArchive_PoolID_path + "/" + fileName

                        # Destination path
                        dest_path = pidFolder+ "/" + fileName
                        
                        # Copy the file
                        try:
                            #print(f"\t Copying from\t {src_path}\n\t to\t {dest_path}") # debug
                            shutil.copy2(src_path, dest_path)
                            print(f"\t Copy Success: {fileName}") # debug
                        except Exception as e:
                            print(f"\t Copy Failed!:{e}")
                        
                        # add a link to the PoolID.html that opens PoolDate.pdf
                        pid_html_content += f'    <li><a target="_blank" href="{fileName}">{fileName}</a></li>\n'
                        
                    except IndexError as e:
                        # print(f"{e}") # debug
                        break
                    except Exception as e:
                        print(style.RED + f"!--ERROR:{e}" + style.RESET)
                        break
                print(f"\t {file} pdfs transferred")
            
            # save PoolID.html
            pid_html_content += "</ul>\n</body>\n</html>"
            with open(pid_html_path, "w") as file:
                file.write(pid_html_content)
            # print(f"\t {pid_html_name} has been generated.") # debug
            
            # add a link to the CustomerID.html that leads to PoolID.html
            cid_html_content += f'    <li><a target="right" href="{PoolID}/{pid_html_name}">{PoolID}</a></li>\n'

        # save CustomerID.html
        cid_html_content += "</ul>\n</body>\n</html>"
        with open(cid_html_path, "w") as file:
            file.write(cid_html_content)
        
        # print(f" {cid_html_name} has been generated.") # debug
        
        
        # add a link to AllCustomers.html that leads to CustomerID.html
        all_cids_html_content += f'    <li><a target="bottom" href="{CustomerID}/{cid_html_name}">{CustomerID}</a> {CustomerName}</li>\n'
    
    
    # Loop that Creates a Branch Page each cycle:
    
    for num in range(0,len(ArrayT_ID)):
        # This loop iterates through the Array of T_IDs,
        # creating a branch page for each one
        
        # Create CustomerID.html in Customers Folder
        CustomerID = ArrayT_ID[num]
        tid_html_name = CustomerID +".html"
        tid_html_path = output_dir + "/" + CustomerID + "/" + tid_html_name
        tid_html_content= f"<!DOCTYPE html>\n<html>\n<head>\n<title>{CustomerID}</title>\n</head>\n<body>\n<h2>{CustomerID}</h2>\n<h3>Locations:</h3>\n<ul>\n"
            
        # iterate through tripleArray
        for num2 in range(0,len(tripleArray[num])):
            # This loop iterates through the array of Names
            # that all share a T_ID
            # creating a link in the branch page to each one
            # they already exist and already link back to the branch page
            
            #populating each tid_html_content with each element inside
            tempID = CustomerID+"_"+str(num2+1)
            tempstring = f'    <li><a target="bottom" href="../{tempID}/{tempID}.html">{tempID}</a> {tripleArray[num][num2]}</li>\n'
            tid_html_content += tempstring
        
        # Cap off and save the Branch Page as an html file in Customers folder
        tid_html_content += "</ul>\n</body>\n</html>"
        with open(tid_html_path, "w") as file:
            file.write(tid_html_content)
    
    # save AllCustomers.html in Customers Folder
    all_cids_html_content += "</ul>\n</body>\n</html>"
    with open(all_cids_html_path, "w") as file:
        file.write(all_cids_html_content)

    # close CustomersPools excel document
    workbook.close()
    
    iterateStats(WEB_ARCHIVE_PATH+"/statsFolder")

    # Prompt the user to press Enter before closing
    input(style.GREEN + "\n\t\t Press Enter to close..." + style.RESET)