#### Created by David N Preiss

### TABLE OF CONTENTS
"""
# Intro
# Import Statements
# Function Definitions
# Pseudocode
# Hard-coded Global Values
# Main Code
# Outro
"""

### Intro
"""
# This file iterates through an xlsx file to navigate the dumpfolder
# To find pool pdfs and organize them in a folder structure
# and then generate html files that navigate that structure
# refer to the pseudocode for details
"""

### Import Statements
if True:
    import shutil
    import os
    import subprocess
    # System call
    os.system("")

    # Class of different styles
    class style():
        BLACK = '\033[30m'
        RED = '\033[31m'
        GREEN = '\033[32m'
        YELLOW = '\033[33m'
        BLUE = '\033[34m'
        MAGENTA = '\033[35m'
        CYAN = '\033[36m'
        WHITE = '\033[37m'
        UNDERLINE = '\033[4m'
        RESET = '\033[0m'

    try:
        import openpyxl
    except ImportError as e:
        print(style.RED + f"!--ERROR:{e}\nopenpyxl is not installed. Installing..." + style.RESET)
        subprocess.check_call(["pip", "install", "openpyxl"])
        print("Installation complete. You can now run the script.")
        # exit()

### Function Definitions

### Hard-coded Global Values
if True:
    MAX_WEEKS = 52
    DUMP_FOLDER_PATH = "../BNRPools/print_populated_pools/Dump Folder"
    CUSTOMERS_POOLS_XLSX = "Book1.xlsx"
    C_ID_COL = 2
    C_ID_START_ROW = 2
    C_ID_END_ROW = 1000
    P_ID_START_COL = 3
    P_ID_END_COL = 1000

### Pseudocode
    """
    # Create Customers Folder

    # Open CustomersPools excel document

    #TODO Create AllCustomers.html in Customers Folder

    # iterate through the CustomerID column in the xlsx
    # for each CustomerID
        # create a CustomerID folder in the Customers folder
        # create a CustomerID.html file in that CustomerID folder
        # iterate through the row of the CustomerID in the xlsx
        # for each PoolID
            # create a PoolID folder in the CustomerID folder
            # inside the PoolID folder
                # create a PoolID.html file
                # for each of the last 52 pdfs of PoolID
                    # import and rename the pdf into the PoolID folder
                    # add a link to the PoolID.html that opens PoolDate.pdf
                # save PoolID.html
            # add a link to the CustomerID.html that leads to PoolID.html
        # save CustomerID.html
        #TODO add a link to AllCustomers.html that leads to CustomerID.html

    #TODO save AllCustomers.html in Customers Folder
    # close CustomersPools excel document
    """

### Main Code
if True:

    # Open CustomersPools excel document
    workbook = openpyxl.load_workbook(CUSTOMERS_POOLS_XLSX)
    worksheet = workbook.active


    # Create Customers Folder
    output_dir = "./Customers"
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    os.makedirs(output_dir, exist_ok=True)
    # print(f"Created output_dir:{output_dir}") # debug
    # output_dir = os.path.abspath(output_dir) # debug


    #TODO Create AllCustomers.html in Customers Folder

    # iterate through the CustomerID column in the xlsx
    for row in range(C_ID_START_ROW, C_ID_END_ROW):
        CustomerID = worksheet.cell(row=row, column=C_ID_COL).value
        if CustomerID is None:
            break
        print(f"\n CustomerID: {CustomerID}")
        # for each CustomerID
        # create a CustomerID folder in the Customers folder
        cidFolder = output_dir + "/"+CustomerID
        os.makedirs(cidFolder, exist_ok=True)
        # print(f" Created cidFolder:{cidFolder}") # debug
        
        # create a CustomerID.html file in that CustomerID folder
        cid_html_name = CustomerID+".html"
        cid_html_path = cidFolder + "/" + cid_html_name
        cid_html_content= f"<!DOCTYPE html>\n<html>\n<head>\n<title>{CustomerID}</title>\n</head>\n<body>\n<h2>{CustomerID}</h2>\n<ul>\n"
        # iterate through the row of the CustomerID in the xlsx
        for col in range(P_ID_START_COL, P_ID_END_COL):
            PoolID = worksheet.cell(row=row, column=col).value
            if PoolID is None:
                break
            print(f"   PoolID: {PoolID}")
            # for each PoolID
            # create a PoolID folder in the CustomerID folder
            pidFolder = cidFolder+"/"+PoolID
            os.makedirs(pidFolder, exist_ok=True)
            # print(f"\t Created pidFolder:{pidFolder}") # debug
            
            # create a PoolID.html file in that PoolID folder
            pid_html_name = PoolID+".html"
            pid_html_path = pidFolder + "/" + pid_html_name
            pid_html_content= f"<!DOCTYPE html>\n<html>\n<head>\n<title>{PoolID}</title>\n</head>\n<body>\n<h2>{PoolID}</h2>\n<ul>\n"
            
            # iterate through the PoolID folder in the dumpfile
            Dump_Pool_ID_path = DUMP_FOLDER_PATH+"/"+str(PoolID)
            fnames = os.listdir(Dump_Pool_ID_path)
            for file in range(0,MAX_WEEKS):
                try:
                    fileName = os.path.basename(fnames[file])
                    # print(f"\t{fileName}") #debug
                    # import and rename the pdf into the PoolID folder
                    # Source path
                    src_path = Dump_Pool_ID_path + "/" + fileName

                    # Destination path
                    dest_path = pidFolder+ "/" + fileName
                    
                    # Copy the file
                    try:
                        # print(f"\t Copying from\t {src_path}\n\t to\t {dest_path}") # debug
                        shutil.copy2(src_path, dest_path)
                        # print(f"\t Copy Success: {fileName}") # debug
                    except Exception as e:
                        print(f"\t Copy Failed!:{e}")
                    
                    # add a link to the PoolID.html that opens PoolDate.pdf
                    pid_html_content += f'    <li><a target="_blank" href="{fileName}">{fileName}</a></li>\n'
                    
                except IndexError as e:
                    # print(f"{e}") # debug
                    break
                except Exception as e:
                    print(style.RED + f"!--ERROR:{e}" + style.RESET)
                    break
            print(f"\t {file} pdfs transferred")
            
            # save PoolID.html
            pid_html_content += "</ul>\n</body>\n</html>"
            with open(pid_html_path, "w") as file:
                file.write(pid_html_content)
            # print(f"\t {pid_html_name} has been generated.") # debug
            
            # add a link to the CustomerID.html that leads to PoolID.html
            cid_html_content += f'    <li><a target="right" href="{PoolID}/{pid_html_name}">{PoolID}</a></li>\n'

        # save CustomerID.html
        cid_html_content += "</ul>\n</body>\n</html>"
        with open(cid_html_path, "w") as file:
            file.write(cid_html_content)
        # print(f" {cid_html_name} has been generated.") # debug
        
        #TODO add a link to AllCustomers.html that leads to CustomerID.html

    #TODO save AllCustomers.html in Customers Folder

    # close CustomersPools excel document
    workbook.close()

    # Prompt the user to press Enter before closing
    input(style.GREEN + "\n\t\t Press Enter to close..." + style.RESET)